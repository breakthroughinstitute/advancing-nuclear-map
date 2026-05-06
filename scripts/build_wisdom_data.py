#!/usr/bin/env python3
"""
build_wisdom_data.py — Regenerate all WISdom-derived data/*.json files
=======================================================================
Replaces migrate_data.py (which reads from frozen _archive/*.js snapshots).
Reads directly from WISdom Google Drive files.

Data sources
------------
A. nuclear_interactive_tool_data.xlsx
     → plant_meta.json      (plant names + state mapping)

B. WISdom siting CSVs  (InstalledCapacity_locations_*.csv)
     → plants.json          (nuclear/coal/ARTES/etc. sites per scenario/year)

C. Coal conversion CSVs  (sitesConverted_2_*.csv)
     → coal_sites.json      (coal plant conversion opportunities)

D. WISdom summary xlsx  (WISdomP_Full_Run_BTI_*.xlsx, one per scenario)
     → state_capacity.json  (state-level installed capacity by tech/year)
     → state_scenarios.json (reactors, jobs, capital investment by state/year)
     → nuclear_detail.json  (per-tech ROI, investment, revenue, jobs by state/year)
     → dispatch.json        (hourly summer-week generation — July 1-7, 2050, by tech)
     → national.json        (national capacity, capital costs, emissions)

E. retailRates_allScenarios.csv
     → national.json        (retail_price: ¢/kWh by scenario/year)

F. Financial CSVs  (ROI_*, costOfInvestment_*, revenues_*)
     → nuclear_detail.json  (roi, inv, rev fields)

G. Data - AS 9-30-25 - v1 (1).xlsx
     → state_investment.json (state capital investment $B by scenario/year)
     → national.json        (adv_nuke_jobs: advanced nuclear jobs by scenario/year)

Does NOT touch (have their own dedicated scripts):
    data/uprate.json, data/orphan.json, data/state_policy.json, data/nrc_pipeline.json

Does NOT regenerate (static):
    data/geodata.json  — standard US states GeoJSON; never changes

Usage:
    pip install openpyxl pandas
    python3 scripts/build_wisdom_data.py

Google Drive is auto-detected from ~/Library/CloudStorage/GoogleDrive-*/.
"""

import csv
import glob
import json
import math
import os

import openpyxl
import pandas as pd

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR  = os.path.join(REPO_ROOT, "data")

# ─────────────────────────────────────────────────────────────────────────────
# Google Drive root auto-detection
# ─────────────────────────────────────────────────────────────────────────────

def find_gdrive_root():
    env = os.environ.get("GDRIVE_ROOT")
    if env and os.path.isdir(os.path.join(env, "Shared drives")):
        return env
    for c in glob.glob(os.path.expanduser("~/Library/CloudStorage/GoogleDrive-*/")):
        if os.path.isdir(os.path.join(c, "Shared drives")):
            return c.rstrip("/")
    for path in ["/Volumes/GoogleDrive", os.path.expanduser("~/Google Drive")]:
        if os.path.isdir(os.path.join(path, "Shared drives")):
            return path
    raise RuntimeError(
        "Could not find Google Drive mount. "
        "Set GDRIVE_ROOT env var to the folder containing 'Shared drives/'.\n"
        "  Example: export GDRIVE_ROOT=\"$HOME/Library/CloudStorage/GoogleDrive-you@example.com\""
    )

GDRIVE_ROOT  = find_gdrive_root()
NUCLEAR_BASE = os.path.join(
    GDRIVE_ROOT,
    "Shared drives/Nuclear/Projects/2022/Advancing Nuclear Report (2022)"
)
print(f"Google Drive root: {GDRIVE_ROOT}")

XLSX_INTERACTIVE   = os.path.join(NUCLEAR_BASE, "Interactive_Tool/nuclear_interactive_tool_data.xlsx")
SITING_IMAGES_BASE = os.path.join(NUCLEAR_BASE, "WISdom outputs/Siting/Siting Images")
SITING_CONV_BASE   = os.path.join(NUCLEAR_BASE, "WISdom outputs/Siting")
SPREADSHEETS_BASE  = os.path.join(NUCLEAR_BASE, "WISdom outputs/Spreadsheets")
FINANCIAL_BASE     = os.path.join(NUCLEAR_BASE, "WISdom outputs/Financial")
RETAIL_CSV         = os.path.join(NUCLEAR_BASE, "WISdom outputs/retailRates_allScenarios.csv")
AS930_XLSX         = os.path.join(NUCLEAR_BASE, "Data sheets for charts/Data - AS 9-30-25 - v1 (1).xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

SCENARIOS = ["LowCost LowLR", "LowCost HighLR", "HighCost LowLR", "HighCost HighLR"]
YEARS     = ["2020", "2025", "2030", "2035", "2040", "2045", "2050"]
N_TO_YEAR = {1: "2020", 2: "2025", 3: "2030", 4: "2035", 5: "2040", 6: "2045", 7: "2050"}

SCENARIO_FILE_KEY = {
    "LowCost LowLR":  "LowCost_LowLR",
    "LowCost HighLR": "LowCost_HighLR",
    "HighCost LowLR": "HighCost_LowLR",
    "HighCost HighLR":"HighCost_HighLR",
}

# Scenario name mapping used in AS-9.30.v1 xlsx (Costs × Learning)
AS930_SCENARIO_MAP = {
    ("Lower", "Low"):    "LowCost LowLR",
    ("Lower", "High"):   "LowCost HighLR",
    ("Lower", "Higher"): "LowCost HighLR",   # "Higher" = HighLR
    ("Upper", "Low"):    "HighCost LowLR",
    ("Upper", "High"):   "HighCost HighLR",
    ("Upper", "Higher"): "HighCost HighLR",
}

STATE_CAP_TECH_COLS = {
    "Coal", "NG CC", "NG GT", "UtilStorage", "Nuclear", "Hydro",
    "Wind", "Offshore", "DPV", "UPV", "CSP", "Geo/Bio", "CCS",
    "SMR", "ARTES", "HTGR",
}

TECH_COL_MAP = {
    "Coal":    "Coal",   "NGCC":     "NG CC",   "NGCT":    "NG GT",
    "Storage": "UtilStorage",        "Nuclear":  "Nuclear", "Hydro":   "Hydro",
    "Wind":    "Wind",   "Offshore": "Offshore", "DPV":     "DPV",
    "UPV":     "UPV",    "CSP":      "CSP",      "Geo/Bio": "Geo/Bio",
    "CCS":     "CCS",    "SMR":      "SMR",      "ARTES":   "ARTES",
    "HTGR":    "HTGR",
}
TECH_COLS  = list(TECH_COL_MAP.keys())
TECHS_CONV = ["ARTES", "SMR", "HTGR"]

# Dispatch tech column → JSON key (matches dashboard.js DISP usage)
DISPATCH_TECH_MAP = {
    "Coal": "Coal", "NG CC": "NG CC", "NG GT": "NG GT",
    "Storage": "Storage", "Nuclear": "Nuclear", "Hydro": "Hydro",
    "Wind": "Wind", "Offshore": "Offshore", "DPV": "DPV", "UPV": "UPV",
    "CSP": "CSP", "Geo/Bio": "Geo/Bio", "CCS": "CCS",
    "SMR": "SMR", "ARTES": "ARTES", "HTGR": "HTGR",
}
DISPATCH_TECH_COLS = list(DISPATCH_TECH_MAP.keys())

# Summer week: first full week of July (July 1-7, 168 hours)
# Extracted from the 2050 Aggregated Dispatch sheet (most built-out scenario year)
SUMMER_WEEK_MONTH = "07"  # July
SUMMER_WEEK_START_DAY = "01"
SUMMER_WEEK_HOURS = 168  # 7 days × 24 hours

STATE_ABBREV = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY", "District of Columbia": "DC",
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def clean_val(v):
    if v is None:
        return None
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
    except (TypeError, ValueError):
        pass
    return v

def parse_year(v):
    """Convert a cell value to a 4-digit year string, or None if not a valid year."""
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return None

def safe_float(v, default=0.0):
    v = clean_val(v)
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default

def safe_int(v):
    v = clean_val(v)
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def fmt_key(lat, lon):
    return f"{float(lat):.3f},{float(lon):.3f}"

def write_json(filename, data, compact=True):
    path = os.path.join(DATA_DIR, filename)
    sep = (",", ":") if compact else (",", ": ")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, separators=sep, ensure_ascii=False)
    size_kb = os.path.getsize(path) / 1024
    print(f"  {filename:35s}  {size_kb:7.0f} KB")

def load_wisdom_xlsx(scenario):
    fkey = SCENARIO_FILE_KEY[scenario]
    path = os.path.join(SPREADSHEETS_BASE, f"WISdomP_Full_Run_BTI_{fkey}_RCP00.xlsx")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Missing WISdom xlsx: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

def parse_sub_table(ws_rows, start_col, years=YEARS):
    """Parse a state×year sub-table from a compound sheet given its start column."""
    data = {}
    for row in ws_rows[2:]:  # skip 2 header rows
        state = row[start_col] if start_col < len(row) else None
        state = clean_val(state)
        if state is None:
            continue
        state = str(state).strip()
        state_data = {}
        for col_offset, year in enumerate(years, start=1):
            idx = start_col + col_offset
            val = row[idx] if idx < len(row) else None
            state_data[year] = safe_float(val)
        data[state] = state_data
    return data


# ─────────────────────────────────────────────────────────────────────────────
# A. plant_meta.json  (PN + PSTATE)
# Source: nuclear_interactive_tool_data.xlsx → "Plant Names" sheet
# ─────────────────────────────────────────────────────────────────────────────

def build_plant_meta():
    print("  Reading Plant Names sheet...")
    df = pd.read_excel(XLSX_INTERACTIVE, sheet_name="Plant Names")
    PN, PSTATE = {}, {}
    for _, row in df.iterrows():
        lat, lon = clean_val(row["Lat"]), clean_val(row["Lon"])
        name = clean_val(row["Plant Name"])
        if lat is None or lon is None or name is None:
            continue
        key = fmt_key(lat, lon)
        PN[key] = str(name)
        state = row.get("State")
        if clean_val(state) is not None:
            PSTATE[key] = str(state)
    print(f"    PN: {len(PN)}, PSTATE: {len(PSTATE)}")
    return {"names": PN, "states": PSTATE}


# ─────────────────────────────────────────────────────────────────────────────
# B. plants.json  (DATA.g — WISdom plant siting)
# Source: InstalledCapacity_locations_WISdomP_Full_Run_BTI_*.csv
# ─────────────────────────────────────────────────────────────────────────────

def build_plants():
    NUCLEAR_DOMS = {"Nuclear", "SMR", "ARTES", "HTGR"}
    DATA_g = {}
    for scenario in SCENARIOS:
        folder = os.path.join(SITING_IMAGES_BASE, scenario)
        scenario_data = {}
        for n, year in N_TO_YEAR.items():
            fkey = SCENARIO_FILE_KEY[scenario]
            fname = f"{n}-InstalledCapacity_locations_WISdomP_Full_Run_BTI_{fkey}_RCP00.csv"
            fpath = os.path.join(folder, fname)
            if not os.path.exists(fpath):
                print(f"    WARNING: missing {fpath}")
                scenario_data[year] = []
                continue
            df = pd.read_csv(fpath)
            tech_sum = df[TECH_COLS].sum(axis=1)
            df = df[tech_sum > 0].copy()
            df["_total"] = df[TECH_COLS].sum(axis=1)
            df["_dom"] = df[TECH_COLS].apply(
                lambda r: TECH_COL_MAP[max(TECH_COLS, key=lambda c: r[c])], axis=1
            )
            df = df.nlargest(2000, "_total")
            rows = []
            for _, row in df.iterrows():
                lat  = round(float(row["Latitude"]), 4)
                lon  = round(float(row["Longitude"]), 4)
                total = round(float(row["_total"]), 1)
                dom  = row["_dom"]
                techs = {
                    TECH_COL_MAP[c]: round(float(row[c]), 1)
                    for c in TECH_COLS if float(row[c]) > 0
                }
                rows.append([lat, lon, total, dom, techs])
            scenario_data[year] = rows
        DATA_g[scenario] = scenario_data
        total = sum(len(v) for v in scenario_data.values())
        print(f"    DATA.g [{scenario}]: {total} rows across {len(scenario_data)} years")
    return DATA_g


# ─────────────────────────────────────────────────────────────────────────────
# C. coal_sites.json  (DATA.c — coal conversion sites)
# Source: sitesConverted_2_*.csv
# ─────────────────────────────────────────────────────────────────────────────

def build_coal_sites():
    DATA_c = {}
    for scenario in SCENARIOS:
        plants = {}
        for tech in TECHS_CONV:
            fname = f"sitesConverted_2_{tech}{scenario}.csv"
            fpath = os.path.join(SITING_CONV_BASE, fname)
            if not os.path.exists(fpath):
                print(f"    WARNING: missing {fpath}")
                continue
            df = pd.read_csv(fpath)
            cap_col = f"{tech} Capacity"
            if cap_col not in df.columns:
                alt = [c for c in df.columns if tech in c]
                cap_col = alt[0] if alt else None
            if cap_col is None:
                continue
            for _, row in df.iterrows():
                pid = clean_val(row.get("plant_id_eia"))
                if pid is None:
                    continue
                pid = int(float(pid))
                lat  = clean_val(row.get("latitude"))
                lon  = clean_val(row.get("longitude"))
                name = clean_val(row.get("plant_name_eia"))
                state = clean_val(row.get("State"))
                orig_cap = clean_val(row.get("nameplate_cap"))
                new_cap  = clean_val(row.get(cap_col))
                orig_tech = clean_val(row.get("technology_description"))
                decom = safe_int(row.get("Decommision Year"))
                conv  = safe_int(row.get("Convertion Year"))
                new_cap_v = float(new_cap) if new_cap is not None else 0.0
                record = [
                    float(lat)  if lat  is not None else None,
                    float(lon)  if lon  is not None else None,
                    str(name)   if name is not None else "",
                    str(state)  if state is not None else "",
                    round(float(orig_cap) if orig_cap is not None else 0.0, 2),
                    round(new_cap_v, 2),
                    str(orig_tech) if orig_tech is not None else "",
                    tech, decom, conv,
                ]
                if pid not in plants or plants[pid][5] < new_cap_v:
                    plants[pid] = record
        DATA_c[scenario] = list(plants.values())
        print(f"    DATA.c [{scenario}]: {len(DATA_c[scenario])} plants")
    return DATA_c


# ─────────────────────────────────────────────────────────────────────────────
# D-1. state_capacity.json  (SC)
# Source: WISdom xlsx → "State Capacities YYYY" sheets
# ─────────────────────────────────────────────────────────────────────────────

def build_state_capacity():
    SC = {}
    for scenario in SCENARIOS:
        print(f"    Parsing state_capacity [{scenario}]...")
        wb = load_wisdom_xlsx(scenario)
        state_cap = {}
        for year in YEARS:
            sheet = f"State Capacities {year}"
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            headers = {
                c: str(rows[1][c]).strip()
                for c in range(len(rows[1]))
                if clean_val(rows[1][c]) is not None
                   and str(rows[1][c]).strip() in STATE_CAP_TECH_COLS
            }
            for row in rows[2:]:
                state = clean_val(row[0])
                if state is None:
                    continue
                state = str(state).strip()
                state_cap.setdefault(state, {})[year] = {}
                for col_idx, tech in headers.items():
                    val = clean_val(row[col_idx]) if col_idx < len(row) else None
                    try:
                        mw = float(val) if val is not None else 0.0
                        if mw > 0:
                            state_cap[state][year][tech] = round(mw, 1)
                    except (TypeError, ValueError):
                        pass

        # Parse jobs from " Jobs" sheet
        ws_jobs = wb[" Jobs"]
        rows_j  = list(ws_jobs.iter_rows(values_only=True))
        nuke_job_types = {"Nuclear Jobs", "ARTES Jobs", "SMR Jobs", "HTGR Jobs"}
        # Find sub-table start columns from row 0 (job type headers)
        job_subtables = {}
        for col_idx, val in enumerate(rows_j[0]):
            if clean_val(val) is not None:
                job_subtables[str(val).strip()] = col_idx

        jobs_by_state = {}
        for jtype, start_col in job_subtables.items():
            tbl = parse_sub_table(rows_j, start_col)
            for state, yr_data in tbl.items():
                jobs_by_state.setdefault(state, {})
                for year, val in yr_data.items():
                    jobs_by_state[state].setdefault(year, {"jobsNuke": 0.0, "jobsTotal": 0.0})
                    if jtype in nuke_job_types:
                        jobs_by_state[state][year]["jobsNuke"] += val
                    jobs_by_state[state][year]["jobsTotal"] += val

        # Parse capital investment
        cap_by_state = {}
        for tech in ["ARTES", "SMR", "HTGR"]:
            ws_cap = wb[f"Capital Investment {tech}"]
            rows_c = list(ws_cap.iter_rows(values_only=True))
            for row in rows_c[2:]:
                state = clean_val(row[0])
                if state is None:
                    continue
                state = str(state).strip()
                cap_by_state.setdefault(state, {})
                for col_offset, year in enumerate(YEARS, start=1):
                    val = row[col_offset] if col_offset < len(row) else None
                    cap_by_state[state][year] = (
                        cap_by_state[state].get(year, 0.0) + safe_float(val)
                    )

        # Merge into SC
        all_states = set(state_cap) | set(jobs_by_state) | set(cap_by_state)
        scenario_sc = {}
        for state in sorted(all_states):
            state_data = {}
            for year in YEARS:
                entry = dict(state_cap.get(state, {}).get(year, {}))
                cap_nuke = cap_by_state.get(state, {}).get(year, 0.0)
                j = jobs_by_state.get(state, {}).get(year, {})
                entry["capNuke"]   = cap_nuke
                entry["capTotal"]  = cap_nuke
                entry["jobsNuke"]  = j.get("jobsNuke", 0.0)
                entry["jobsTotal"] = j.get("jobsTotal", 0.0)
                state_data[year] = entry
            scenario_sc[state] = state_data
        SC[scenario] = scenario_sc
        print(f"      {len(scenario_sc)} states")
    return SC


# ─────────────────────────────────────────────────────────────────────────────
# D-2. state_scenarios.json  (SD)
# Source: WISdom xlsx → "Reactors by State" + " Jobs" + "Capital Investment *"
# ─────────────────────────────────────────────────────────────────────────────

def build_state_scenarios():
    SD = {}
    for scenario in SCENARIOS:
        print(f"    Parsing state_scenarios [{scenario}]...")
        wb = load_wisdom_xlsx(scenario)

        # Reactor counts by type
        ws_r  = wb["Reactors by State"]
        rows_r = list(ws_r.iter_rows(values_only=True))
        smr   = parse_sub_table(rows_r, 0)
        artes = parse_sub_table(rows_r, 9)
        htgr  = parse_sub_table(rows_r, 18)
        trad  = parse_sub_table(rows_r, 27)

        # Jobs
        ws_j   = wb[" Jobs"]
        rows_j = list(ws_j.iter_rows(values_only=True))
        nuke_job_types = {"Nuclear Jobs", "ARTES Jobs", "SMR Jobs", "HTGR Jobs"}
        job_subtables = {}
        for col_idx, val in enumerate(rows_j[0]):
            if clean_val(val) is not None:
                job_subtables[str(val).strip()] = col_idx

        jobs_nuke = {}
        jobs_total = {}
        for jtype, start_col in job_subtables.items():
            tbl = parse_sub_table(rows_j, start_col)
            for state, yr_data in tbl.items():
                for year, val in yr_data.items():
                    if jtype in nuke_job_types:
                        jobs_nuke.setdefault(state, {})[year] = (
                            jobs_nuke.get(state, {}).get(year, 0.0) + val
                        )
                    jobs_total.setdefault(state, {})[year] = (
                        jobs_total.get(state, {}).get(year, 0.0) + val
                    )

        # Capital investment (sum ARTES + SMR + HTGR)
        cap_nuke = {}
        for tech in ["ARTES", "SMR", "HTGR"]:
            ws_c  = wb[f"Capital Investment {tech}"]
            rows_c = list(ws_c.iter_rows(values_only=True))
            for row in rows_c[2:]:
                state = clean_val(row[0])
                if state is None:
                    continue
                state = str(state).strip()
                cap_nuke.setdefault(state, {})
                for col_offset, year in enumerate(YEARS, start=1):
                    val = row[col_offset] if col_offset < len(row) else None
                    cap_nuke[state][year] = (
                        cap_nuke[state].get(year, 0.0) + safe_float(val)
                    )

        all_states = set(smr) | set(artes) | set(htgr) | set(trad)
        scenario_sd = {}
        for state in sorted(all_states):
            abbrev = STATE_ABBREV.get(state, state)
            state_data = {}
            for year in YEARS:
                s = safe_float(smr.get(state, {}).get(year))
                a = safe_float(artes.get(state, {}).get(year))
                h = safe_float(htgr.get(state, {}).get(year))
                t = safe_float(trad.get(state, {}).get(year))
                r_total = s + a + h
                cap = cap_nuke.get(state, {}).get(year, 0.0)
                state_data[year] = {
                    "capTotal":  cap,
                    "capNuke":   cap,
                    "reactors":  {"SMR": s, "ARTES": a, "HTGR": h, "Trad": t},
                    "rTotal":    r_total,
                    "jobsNuke":  jobs_nuke.get(state, {}).get(year, 0.0),
                    "jobsTotal": jobs_total.get(state, {}).get(year, 0.0),
                }
            scenario_sd[abbrev] = state_data
        SD[scenario] = scenario_sd
        print(f"      {len(scenario_sd)} states")
    return SD


# ─────────────────────────────────────────────────────────────────────────────
# D-3. nuclear_detail.json  (ND)
# Source: WISdom xlsx " Jobs" + Financial CSVs (ROI/costOfInvestment/revenues)
# ─────────────────────────────────────────────────────────────────────────────

def load_financial_csv(name, scenario):
    """Load a Financial CSV file; return dict {state → {year → float}}."""
    fname = f"{name}_{scenario}.csv"
    fpath = os.path.join(FINANCIAL_BASE, fname)
    if not os.path.exists(fpath):
        return {}
    result = {}
    with open(fpath, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            state = row.get("STATE", "").strip()
            if not state:
                continue
            result[state] = {}
            for year in YEARS:
                result[state][year] = safe_float(row.get(year, ""))
    return result


def build_nuclear_detail():
    ND = {}
    for scenario in SCENARIOS:
        print(f"    Parsing nuclear_detail [{scenario}]...")
        wb = load_wisdom_xlsx(scenario)

        # Per-tech nuclear jobs from " Jobs" sheet
        ws_j   = wb[" Jobs"]
        rows_j = list(ws_j.iter_rows(values_only=True))
        tech_jobs = {}
        for col_idx, val in enumerate(rows_j[0]):
            if clean_val(val) is None:
                continue
            label = str(val).strip()
            for tech in ["SMR", "ARTES", "HTGR"]:
                if label == f"{tech} Jobs":
                    tech_jobs[tech] = parse_sub_table(rows_j, col_idx)

        # Financial CSVs
        roi_data = {t: load_financial_csv(f"ROI_{t.lower()}", scenario) for t in ["ARTES", "SMR", "HTGR"]}
        inv_data = {t: load_financial_csv(f"costOfInvestment_{t.lower()}", scenario) for t in ["ARTES", "SMR", "HTGR"]}
        rev_data_artes = load_financial_csv("revenues_artes", scenario)
        rev_data_smr   = load_financial_csv("revenues_smrr",  scenario)  # note: 'smrr' double-r
        rev_data_htgr  = load_financial_csv("revenues_htgr",  scenario)
        rev_data = {"ARTES": rev_data_artes, "SMR": rev_data_smr, "HTGR": rev_data_htgr}

        # Collect all state names from Financial CSVs
        all_states = set()
        for tech in ["ARTES", "SMR", "HTGR"]:
            all_states.update(roi_data[tech].keys())
            all_states.update(inv_data[tech].keys())

        scenario_nd = {}
        detail_years = [y for y in YEARS if y >= "2030"]  # only years with non-zero data
        for state in sorted(all_states):
            abbrev = STATE_ABBREV.get(state, state)
            state_data = {}
            for year in detail_years:
                tech_entry = {}
                for tech in ["SMR", "ARTES", "HTGR"]:
                    # roi: stored as decimal fraction in CSV × 100 = displayed value
                    roi_raw = roi_data[tech].get(state, {}).get(year, 0.0)
                    inv_val = inv_data[tech].get(state, {}).get(year, 0.0)
                    rev_val = rev_data[tech].get(state, {}).get(year, 0.0)
                    jobs_val = safe_float(
                        tech_jobs.get(tech, {}).get(state, {}).get(year, 0.0)
                    )
                    tech_entry[tech] = {
                        "roi":  round(roi_raw * 100, 2),
                        "inv":  round(inv_val),
                        "rev":  round(rev_val),
                        "jobs": round(jobs_val),
                    }
                state_data[year] = tech_entry
            scenario_nd[abbrev] = state_data
        ND[scenario] = scenario_nd
        print(f"      {len(scenario_nd)} states")
    return ND


# ─────────────────────────────────────────────────────────────────────────────
# D-4. dispatch.json  (DISP — hourly summer week generation)
# Source: WISdom xlsx → "Aggregated Dispatch YYYY" (2050), July 1-7 week
# Each entry = one hour, values in GW (converted from MW ÷ 1000)
# ─────────────────────────────────────────────────────────────────────────────

def build_dispatch():
    DISP = {}
    dispatch_year = "2050"  # use most built-out year for representative dispatch

    for scenario in SCENARIOS:
        print(f"    Parsing dispatch [{scenario}]...")
        wb   = load_wisdom_xlsx(scenario)
        sheet_name = f"Aggregated Dispatch {dispatch_year}"
        ws   = wb[sheet_name]

        # Find column headers from row 2 (row index 1)
        header_row = None
        data_rows  = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if header_row is None:
                if row[0] == "Time" or (isinstance(row[0], str) and "Time" in str(row[0])):
                    header_row = list(row)
                continue
            data_rows.append(row)

        if header_row is None:
            # fallback: row 2 is header
            all_rows = list(ws.iter_rows(values_only=True))
            header_row = list(all_rows[1])
            data_rows  = list(all_rows[2:])

        # Map header names to column indices
        col_idx = {}
        for i, h in enumerate(header_row):
            if h is not None:
                col_idx[str(h).strip()] = i

        # Find start of July 1 week
        prefix = f"{SUMMER_WEEK_MONTH}/{SUMMER_WEEK_START_DAY}/"
        start_row = None
        for i, row in enumerate(data_rows):
            ts = str(row[0]) if row[0] is not None else ""
            if ts.startswith(prefix):
                start_row = i
                break

        if start_row is None:
            print(f"    WARNING: July 1 not found in dispatch {scenario}, using rows 3500-3668")
            start_row = 3500  # fallback: approximate mid-year

        week_rows = data_rows[start_row: start_row + SUMMER_WEEK_HOURS]
        entries = []
        for row in week_rows:
            entry = {}
            for tech, key in DISPATCH_TECH_MAP.items():
                c = col_idx.get(tech)
                if c is not None:
                    mw = safe_float(row[c] if c < len(row) else None)
                    entry[key] = round(mw / 1000, 1)  # MW → GW
            entries.append(entry)

        DISP[scenario] = entries
        print(f"      {len(entries)} hourly entries (July 1-7 week, {dispatch_year})")
    return DISP


# ─────────────────────────────────────────────────────────────────────────────
# D-5. national.json  (NAT, NAT_JOBS, RETAIL, SYSCOST)
# Sources:
#   capacity.cap       ← WISdom xlsx "Capacities" sheet
#   capacity.costs     ← WISdom xlsx "Learned Capital Costs" sheet ($/kW by year)
#   capacity.emissions ← WISdom xlsx " Emissions" sheet
#   adv_nuke_jobs      ← AS-9.30.v1 "Data" sheet (SMR+ARTES+HTGR jobs summed)
#   retail_price       ← retailRates_allScenarios.csv
#   system_cost        ← national total resource cost / estimated generation
# ─────────────────────────────────────────────────────────────────────────────

def build_national():
    NAT    = {}  # capacity.cap
    COSTS  = {}  # capacity.costs (Learned Capital Costs $/kW)
    EMISS  = {}  # capacity.emissions
    NAT_JOBS = {}
    RETAIL = {}
    SYSCOST = {}

    # ── Capacity, costs, emissions (from WISdom xlsx) ─────────────────────
    CAP_TECH_MAP = {
        "Coal": "Coal", "NG CC": "NG CC", "NG GT": "NG GT",
        "UtilStorage": "UtilStorage", "Nuclear": "Nuclear", "Hydro": "Hydro",
        "Wind": "Wind", "Offshore": "Offshore", "DPV": "DPV", "UPV": "UPV",
        "CSP": "CSP", "Geo/Bio": "Geo/Bio", "CCS": "CCS",
        "SMR": "SMR", "ARTES": "ARTES", "HTGR": "HTGR",
    }
    EMISS_COLS = ["CO2", "CO", "SO2", "NOx", "CH4", "N2O", "VOC", "PM2_5", "PM10"]

    for scenario in SCENARIOS:
        print(f"    Parsing national [{scenario}]...")
        wb = load_wisdom_xlsx(scenario)

        # Capacities sheet
        ws_cap   = wb["Capacities"]
        rows_cap = list(ws_cap.iter_rows(values_only=True))
        header   = [str(c).strip() if c else None for c in rows_cap[1]]
        cap_by_year = {}
        for row in rows_cap[2:]:
            year = parse_year(row[0])
            if year is None or year not in YEARS:
                continue
            if year in cap_by_year:
                continue  # first section only — skip Percentage Change and Generation sections
            entry = {}
            for col_idx, col_name in enumerate(header):
                if col_name in CAP_TECH_MAP and col_idx < len(row):
                    mw = safe_float(row[col_idx])
                    if mw > 0:
                        entry[CAP_TECH_MAP[col_name]] = round(mw)
            cap_by_year[year] = entry
        NAT[scenario] = {"cap": cap_by_year}

        # Learned Capital Costs sheet ($/kW by year, varies with learning rate)
        ws_costs   = wb["Learned Capital Costs"]
        rows_costs = list(ws_costs.iter_rows(values_only=True))
        cost_header = [str(c).strip() if c else None for c in rows_costs[1]]
        cost_by_year = {}
        for row in rows_costs[2:]:
            year = parse_year(row[0])
            if year is None or year not in YEARS:
                continue
            entry = {}
            for col_idx, col_name in enumerate(cost_header):
                if col_name and col_idx < len(row):
                    val = safe_float(row[col_idx])
                    if val > 0:
                        entry[col_name] = round(val, 2)
            cost_by_year[year] = entry
        COSTS[scenario] = cost_by_year
        NAT[scenario]["costs"] = cost_by_year

        # Emissions sheet
        ws_em   = wb[" Emissions"]
        rows_em = list(ws_em.iter_rows(values_only=True))
        em_header = [str(c).strip() if c else None for c in rows_em[1]]
        em_by_year = {}
        for row in rows_em[2:]:
            year = parse_year(row[0])
            if year is None or year not in YEARS:
                continue
            if year in em_by_year:
                continue  # first section only — skip Percentage Change section
            entry = {}
            for col_idx, col_name in enumerate(em_header):
                if col_name in EMISS_COLS and col_idx < len(row):
                    val = safe_float(row[col_idx])
                    if val != 0:
                        entry[col_name] = round(val)
            em_by_year[year] = entry
        EMISS[scenario] = em_by_year
        NAT[scenario]["emissions"] = em_by_year

        # System cost: national total resource cost / estimated total generation (MWh)
        # Uses "Total Resource Cost" sheet: sum all state rows for each year
        # Generation estimated from Capacities × 8760 hrs × national avg CF
        ws_trc  = wb["Total Resource Cost"]
        rows_trc = list(ws_trc.iter_rows(values_only=True))
        # headers in row 2, state data from row 3; sum col 1-7 for 2020-2050
        trc_sum = {year: 0.0 for year in YEARS}
        trc_header = [clean_val(c) for c in rows_trc[1]]
        for row in rows_trc[2:]:
            state = clean_val(row[0])
            if state is None or not isinstance(state, str):
                continue
            for col_offset, year in enumerate(YEARS, start=1):
                val = safe_float(row[col_offset] if col_offset < len(row) else None)
                trc_sum[year] += val

        # Estimate total annual generation from capacity (MWh = MW × 8760 × cf)
        # Use a rough national average of 35% CF for the full mix
        AVG_CF = 0.35
        syscost = {}
        for year in YEARS:
            total_cap_mw = sum(cap_by_year.get(year, {}).values())
            total_gen_mwh = total_cap_mw * 8760 * AVG_CF
            if total_gen_mwh > 0:
                syscost[year] = round(trc_sum[year] / total_gen_mwh, 1)
            else:
                syscost[year] = None
        SYSCOST[scenario] = syscost

    # ── adv_nuke_jobs (from AS-9.30.v1 "Data" sheet) ─────────────────────
    print("    Reading adv_nuke_jobs from AS-9.30.v1...")
    wb_as = openpyxl.load_workbook(AS930_XLSX, read_only=True, data_only=True)
    ws_data = wb_as["Data"]
    rows_data = list(ws_data.iter_rows(values_only=True))
    header_d  = [str(c).strip() if c else None for c in rows_data[0]]

    col_year   = header_d.index("Year")
    col_foak   = header_d.index("FOAK")
    col_lr     = header_d.index("Learning Rate")
    col_tech   = header_d.index("Reactor Technology")
    col_jobs_lo = header_d.index("Jobs (Low)")

    adv_techs = {"SMR", "ARTES", "HTGR"}
    for scenario in SCENARIOS:
        NAT_JOBS[scenario] = {year: 0.0 for year in YEARS}

    for row in rows_data[1:]:
        year = parse_year(row[col_year])
        if year is None or year not in YEARS:
            continue
        tech = str(row[col_tech]).strip() if row[col_tech] else ""
        if tech not in adv_techs:
            continue
        foak = str(row[col_foak]).strip() if row[col_foak] else ""
        lr   = str(row[col_lr]).strip()   if row[col_lr]   else ""
        scenario = AS930_SCENARIO_MAP.get((foak, lr))
        if scenario is None:
            continue
        jobs = safe_float(row[col_jobs_lo])
        NAT_JOBS[scenario][year] = NAT_JOBS[scenario].get(year, 0.0) + jobs

    # ── retail_price from retailRates_allScenarios.csv ───────────────────
    print("    Reading retail_price from CSV...")
    with open(RETAIL_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            year = row[""].strip() if "" in row else list(row.values())[0]
            if year not in YEARS:
                continue
            for scen in SCENARIOS:
                RETAIL.setdefault(scen, {})[year] = round(safe_float(row.get(scen)), 2)

    return {
        "capacity":     NAT,
        "adv_nuke_jobs": NAT_JOBS,
        "retail_price": RETAIL,
        "system_cost":  SYSCOST,
    }


# ─────────────────────────────────────────────────────────────────────────────
# E. state_investment.json  (STATE_INV)
# Source: AS-9.30.v1 "State bans" sheet
# Rows: State / Reactor Technology / Costs / Learning / 2020-2050
# Output: {state_abbrev → {scenario → {year → $B}}}
# ─────────────────────────────────────────────────────────────────────────────

def build_state_investment():
    print("    Reading state_investment from AS-9.30.v1...")
    wb_as = openpyxl.load_workbook(AS930_XLSX, read_only=True, data_only=True)
    ws    = wb_as["State bans"]
    rows  = list(ws.iter_rows(values_only=True))

    # Find header row (has "State" in first column)
    header_row_idx = None
    for i, row in enumerate(rows):
        if clean_val(row[0]) == "State":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError("Could not find 'State' header row in 'State bans' sheet")

    header = [str(c).strip() if clean_val(c) is not None else None for c in rows[header_row_idx]]
    year_cols = {year: header.index(year) for year in YEARS if year in header}

    col_state = 0
    col_costs = header.index("Costs")
    col_lr    = header.index("Learning")

    # Accumulate: sum all technologies per state/scenario/year, convert $ → $B
    STATE_INV = {}
    for row in rows[header_row_idx + 1:]:
        state_full = clean_val(row[col_state])
        if state_full is None:
            break  # end of data
        state_full = str(state_full).strip()
        abbrev = STATE_ABBREV.get(state_full)
        if abbrev is None:
            continue
        costs = str(row[col_costs]).strip() if clean_val(row[col_costs]) else ""
        lr    = str(row[col_lr]).strip()    if clean_val(row[col_lr])    else ""
        scenario = AS930_SCENARIO_MAP.get((costs, lr))
        if scenario is None:
            continue
        STATE_INV.setdefault(abbrev, {}).setdefault(scenario, {})
        for year, col_idx in year_cols.items():
            val = safe_float(row[col_idx] if col_idx < len(row) else None)
            cur = STATE_INV[abbrev][scenario].get(year, 0.0)
            STATE_INV[abbrev][scenario][year] = round(cur + val / 1e9, 4)  # $ → $B

    print(f"      {len(STATE_INV)} states")
    return STATE_INV


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=== build_wisdom_data.py ===\n")
    os.makedirs(DATA_DIR, exist_ok=True)

    # These files are NEVER touched:
    PROTECTED = {"uprate.json", "orphan.json", "state_policy.json",
                 "nrc_pipeline.json", "geodata.json"}

    print("A. plant_meta.json")
    write_json("plant_meta.json", build_plant_meta())

    print("\nB. plants.json")
    write_json("plants.json", build_plants())

    print("\nC. coal_sites.json")
    write_json("coal_sites.json", build_coal_sites())

    print("\nD-1. state_capacity.json")
    write_json("state_capacity.json", build_state_capacity())

    print("\nD-2. state_scenarios.json")
    write_json("state_scenarios.json", build_state_scenarios())

    print("\nD-3. nuclear_detail.json")
    write_json("nuclear_detail.json", build_nuclear_detail())

    print("\nD-4. dispatch.json")
    write_json("dispatch.json", build_dispatch())

    print("\nD-5/E. national.json")
    write_json("national.json", build_national())

    print("\nF. state_investment.json")
    write_json("state_investment.json", build_state_investment())

    print(f"\n=== Done ===")
    print(f"Protected (not touched): {', '.join(sorted(PROTECTED))}")


if __name__ == "__main__":
    main()
